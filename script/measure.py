#!/usr/bin/env python3
# coding=utf-8
import os
from openpyxl import Workbook

temp = tuple(os.walk('result'))[0]
datasets_name = tuple(temp[1])
datasets_path = (
    os.path.realpath(os.path.join(temp[0], dataset)) for dataset in temp[1])


def read_rel(dataset_path, algo):
    f = open(os.path.join(dataset_path, algo + '_descending'))
    rel = []
    for line in f.readlines():
        if len(line.split()) == 3:
            rel.append(1 if line.split()[1] == 'R' else 0)
    f.close()
    return tuple(rel)


def calc_P(rel):
    P = [rel[0]]
    for i in range(1, len(rel)):
        P.append(P[-1] + rel[i])
    for i in range(len(P)):
        P[i] /= i + 1
    return tuple(P)


def calc_AP(rel, P, k):
    if k > len(rel):
        return 0.0
    APk = 0.0
    for i in range(k):
        APk += rel[i] * P[i]
    APk /= sum(rel)
    return APk


def calc_RR(rel, k):
    for i in range(k):
        if rel[i] == 1:
            return 1 / (i + 1)
    else:
        return 0.0


def mean(list):
    return sum(list) / len(list)


if __name__ == '__main__':
    HITS_AP3, HITS_AP5, HITS_AP10, HITS_AP20 = [], [], [], []
    SALSA_AP3, SALSA_AP5, SALSA_AP10, SALSA_AP20 = [], [], [], []
    HITS_RR3, HITS_RR5, HITS_RR10, HITS_RR20 = [], [], [], []
    SALSA_RR3, SALSA_RR5, SALSA_RR10, SALSA_RR20 = [], [], [], []

    for dataset_path in datasets_path:
        print('measuring ' + dataset_path)

        rel = read_rel(dataset_path, 'HITS')
        P = calc_P(rel)
        HITS_AP3.append(calc_AP(rel, P, 3))
        HITS_AP5.append(calc_AP(rel, P, 5))
        HITS_AP10.append(calc_AP(rel, P, 10))
        HITS_AP20.append(calc_AP(rel, P, 20))
        HITS_RR3.append(calc_RR(rel, 3))
        HITS_RR5.append(calc_RR(rel, 5))
        HITS_RR10.append(calc_RR(rel, 10))
        HITS_RR20.append(calc_RR(rel, 20))

        rel = read_rel(dataset_path, 'SALSA')
        P = calc_P(rel)
        SALSA_AP3.append(calc_AP(rel, P, 3))
        SALSA_AP5.append(calc_AP(rel, P, 5))
        SALSA_AP10.append(calc_AP(rel, P, 10))
        SALSA_AP20.append(calc_AP(rel, P, 20))
        SALSA_RR3.append(calc_RR(rel, 3))
        SALSA_RR5.append(calc_RR(rel, 5))
        SALSA_RR10.append(calc_RR(rel, 10))
        SALSA_RR20.append(calc_RR(rel, 20))

    AP_wb = Workbook()
    AP_ws = AP_wb.active
    AP_ws.append([''] + list(datasets_name))
    AP_ws.append(['HITS_AP@3'] + list(HITS_AP3))
    AP_ws.append(['HITS_AP@5'] + list(HITS_AP5))
    AP_ws.append(['HITS_AP@10'] + list(HITS_AP10))
    AP_ws.append(['HITS_AP@20'] + list(HITS_AP20))
    AP_ws.append(['SALSA_AP@3'] + list(SALSA_AP3))
    AP_ws.append(['SALSA_AP@5'] + list(SALSA_AP5))
    AP_ws.append(['SALSA_AP@10'] + list(SALSA_AP10))
    AP_ws.append(['SALSA_AP@20'] + list(SALSA_AP20))
    AP_wb.save(os.path.join('result', 'AP.xlsx'))

    MAP_wb = Workbook()
    MAP_ws = MAP_wb.active
    MAP_ws.append(['', 'MAP@3', 'MAP@5', 'MAP@10', 'MAP@20'])
    MAP_ws.append(
        ['HITS'] + [mean(HITS_AP3)] + [mean(HITS_AP5)] + [mean(HITS_AP10)] + [mean(HITS_AP20)])
    MAP_ws.append(
        ['SALSA'] + [mean(SALSA_AP3)] + [mean(SALSA_AP5)] + [mean(SALSA_AP10)] + [mean(SALSA_AP20)])
    MAP_wb.save(os.path.join('result', 'MAP.xlsx'))

    RR_wb = Workbook()
    RR_ws = RR_wb.active
    RR_ws.append([''] + list(datasets_name))
    RR_ws.append(['HITS_RR@3'] + list(HITS_RR3))
    RR_ws.append(['HITS_RR@5'] + list(HITS_RR5))
    RR_ws.append(['HITS_RR@10'] + list(HITS_RR10))
    RR_ws.append(['HITS_RR@20'] + list(HITS_RR20))
    RR_ws.append(['SALSA_RR@3'] + list(SALSA_RR3))
    RR_ws.append(['SALSA_RR@5'] + list(SALSA_RR5))
    RR_ws.append(['SALSA_RR@10'] + list(SALSA_RR10))
    RR_ws.append(['SALSA_RR@20'] + list(SALSA_RR20))
    RR_wb.save(os.path.join('result', 'RR.xlsx'))

    MRR_wb = Workbook()
    MRR_ws = MRR_wb.active
    MRR_ws.append(['', 'MRR@3', 'MRR@5', 'MRR@10', 'MRR@20'])
    MRR_ws.append(
        ['HITS'] + [mean(HITS_RR3)] + [mean(HITS_RR5)] + [mean(HITS_RR10)] + [mean(HITS_RR20)])
    MRR_ws.append(
        ['SALSA'] + [mean(SALSA_RR3)] + [mean(SALSA_RR5)] + [mean(SALSA_RR10)] + [mean(SALSA_RR20)])
    MRR_wb.save(os.path.join('result', 'MRR.xlsx'))
